"""
Utilization Credit Report — Web App
Run with: streamlit run utilization_app_v2.py
"""

import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PS Utilization Credit Report",
    page_icon="📈",
    layout="wide"
)

# ── Constants ─────────────────────────────────────────────────────────────────
NAVY     = "1e2c63"
TEAL     = "4472C4"
WHITE    = "FFFFFF"
LTGRAY   = "F2F2F2"
MID_GRAY = "BDC3C7"

TAG_COLORS = {
    "CREDITED":     "EAF9F1",
    "OVERRUN":      "FDECED",
    "PARTIAL":      "FEF9E7",
    "NON-BILLABLE": "EBEDEE",
    "FF: NO SCOPE DEFINED": "F2F2F2",
}
TAG_BADGE = {
    "CREDITED":     "🟢",
    "OVERRUN":      "🔴",
    "PARTIAL":      "🟡",
    "NON-BILLABLE": "⚫",
    "FF: NO SCOPE DEFINED": "⚪",
}

# ── Stored scope map ──────────────────────────────────────────────────────────
# ── Employees excluded from utilization targets ──────────────────────────────
UTIL_EXEMPT_EMPLOYEES = ["swanson"]  # case-insensitive match

PS_REGION_MAP = {
    "Sydney (NSW)":     "APAC",
    "Manila (PH)":      "APAC",
    "UK":               "EMEA",
    "Spain":            "EMEA",
    "Netherlands":      "EMEA",
    "Northern Ireland": "EMEA",
    "Faroe Islands":    "EMEA",
    "North Macedonia":  "EMEA",
    "Czech Republic":   "EMEA",
    "Serbia":           "EMEA",
    "USA":              "NOAM",
    "Canada":           "NOAM",
}

DEFAULT_SCOPE = {
    "Capture":                 20,
    "Approvals":               17,
    "Reconcile":               17,
    "PSP":                     18,
    "Payments":                30,
    "Reconcile 2.0":           20,
    "CC":                       6,
    "SFTP":                    12,
    "Premium - 10":            10,
    "Premium - 20":            20,
    "E-Invoicing":             15,
    "Capture and E-Invoicing": 30,
    "Additional Subsidiary":    2,
}

# ── Available hours by region/month (2026) ────────────────────────────────────
AVAIL_HOURS = {
    "Spain":            {"2026-01":155.00,"2026-02":155.00,"2026-03":170.50,"2026-04":162.75,"2026-05":155.00,"2026-06":170.50,"2026-07":178.25,"2026-08":162.75,"2026-09":170.50,"2026-10":162.75,"2026-11":155.00,"2026-12":155.00},
    "UK":               {"2026-01":157.50,"2026-02":150.00,"2026-03":165.00,"2026-04":150.00,"2026-05":142.50,"2026-06":165.00,"2026-07":172.50,"2026-08":150.00,"2026-09":165.00,"2026-10":165.00,"2026-11":157.50,"2026-12":157.50},
    "Northern Ireland": {"2026-01":157.50,"2026-02":150.00,"2026-03":157.50,"2026-04":150.00,"2026-05":142.50,"2026-06":165.00,"2026-07":165.00,"2026-08":150.00,"2026-09":165.00,"2026-10":165.00,"2026-11":157.50,"2026-12":157.50},
    "Netherlands":      {"2026-01":168.00,"2026-02":160.00,"2026-03":176.00,"2026-04":152.00,"2026-05":144.00,"2026-06":176.00,"2026-07":184.00,"2026-08":168.00,"2026-09":176.00,"2026-10":176.00,"2026-11":168.00,"2026-12":176.00},
    "Faroe Islands":    {"2026-01":168.00,"2026-02":160.00,"2026-03":176.00,"2026-04":144.00,"2026-05":144.00,"2026-06":176.00,"2026-07":168.00,"2026-08":168.00,"2026-09":176.00,"2026-10":176.00,"2026-11":168.00,"2026-12":168.00},
    "North Macedonia":  {"2026-01":160.00,"2026-02":160.00,"2026-03":168.00,"2026-04":168.00,"2026-05":160.00,"2026-06":176.00,"2026-07":184.00,"2026-08":168.00,"2026-09":168.00,"2026-10":168.00,"2026-11":168.00,"2026-12":176.00},
    "Czech Republic":   {"2026-01":168.00,"2026-02":160.00,"2026-03":176.00,"2026-04":160.00,"2026-05":152.00,"2026-06":176.00,"2026-07":176.00,"2026-08":168.00,"2026-09":168.00,"2026-10":168.00,"2026-11":160.00,"2026-12":168.00},
    "Serbia":           {"2026-01":152.00,"2026-02":152.00,"2026-03":176.00,"2026-04":160.00,"2026-05":160.00,"2026-06":176.00,"2026-07":184.00,"2026-08":168.00,"2026-09":176.00,"2026-10":176.00,"2026-11":160.00,"2026-12":184.00},
    "Canada":           {"2026-01":168.00,"2026-02":160.00,"2026-03":176.00,"2026-04":168.00,"2026-05":160.00,"2026-06":176.00,"2026-07":176.00,"2026-08":160.00,"2026-09":160.00,"2026-10":168.00,"2026-11":160.00,"2026-12":168.00},
    "USA":              {"2026-01":160.00,"2026-02":152.00,"2026-03":176.00,"2026-04":176.00,"2026-05":160.00,"2026-06":168.00,"2026-07":176.00,"2026-08":168.00,"2026-09":168.00,"2026-10":168.00,"2026-11":152.00,"2026-12":176.00},
    "Sydney (NSW)":     {"2026-01":152.00,"2026-02":152.00,"2026-03":167.20,"2026-04":144.40,"2026-05":159.60,"2026-06":159.60,"2026-07":174.80,"2026-08":152.00,"2026-09":167.20,"2026-10":159.60,"2026-11":159.60,"2026-12":159.60},
    "Manila (PH)":      {"2026-01":168.00,"2026-02":152.00,"2026-03":176.00,"2026-04":152.00,"2026-05":160.00,"2026-06":168.00,"2026-07":184.00,"2026-08":152.00,"2026-09":176.00,"2026-10":176.00,"2026-11":152.00,"2026-12":144.00},
}

# Fixed fee task keywords (Case/Task/Event column)
FF_TASKS = ["Configuration", "Enablement", "Training", "Post Go-live", "Project Management"]

def get_avail_hours(region, period):
    """Look up available hours for a region/period. Returns None if not found."""
    region_clean = str(region).strip()
    # Try exact match first, then case-insensitive
    for r, months in AVAIL_HOURS.items():
        if r.lower() == region_clean.lower():
            return months.get(str(period), None)
    return None

def match_ff_task(task_val):
    """Match a task value to one of the 4 standard FF task categories."""
    t = str(task_val).strip().lower()
    if "config" in t:             return "Configuration"
    if "enabl" in t or "train" in t: return "Enablement/Training"
    if "post" in t or "go-live" in t or "golive" in t: return "Post Go-live Support"
    if "project mgmt" in t or "project management" in t or "pm" == t: return "Project Management"
    return None

# ── Excel helpers ─────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color=MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_fill(hex_color):  return PatternFill("solid", fgColor=hex_color)
def row_fill(hex_color):  return PatternFill("solid", fgColor=hex_color)

GROUP_COLORS = ["EEF2FB", "FFFFFF"]  # alternating soft blue/white for grouped rows

def group_bg(value, prev_value, group_idx):
    """Return bg color and updated group index for grouped first-column display."""
    if value != prev_value:
        group_idx = 1 - group_idx  # toggle group
    return GROUP_COLORS[group_idx], group_idx

def style_header(ws, row, headers, fill_color=NAVY):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = Font(name="Manrope", bold=True, color=WHITE, size=10)
        c.fill      = hdr_fill(fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[row].height = 22

def style_cell(cell, bg, fmt=None, bold=False, align="left"):
    cell.fill      = row_fill(bg)
    cell.font      = Font(name="Manrope", size=10, bold=bold)
    cell.border    = thin_border()
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

def write_title(ws, title, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Manrope", bold=True, size=14, color=WHITE)
    c.fill      = hdr_fill(NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

# ── Column detection ──────────────────────────────────────────────────────────
def auto_detect_columns(df):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    checks = {
        "employee":      ["employee", "employee name", "name", "resource"],
        "project":       ["project", "project name", "job"],
        "project_type":  ["project type", "type", "project_type"],
        "date":          ["date", "time entry date", "entry date", "work date"],
        "hours":         ["hours", "duration", "hours logged", "time", "qty"],
        "approval":      ["approval status", "approval", "status"],
        "task":          ["case/task/event", "task", "case", "event", "memo"],
        "non_billable":  ["non-billable", "non billable", "nonbillable",
                          "non_billable", "is non billable"],
        "billing_type":  ["billing type", "billing_type", "bill type", "billtype"],
        "hours_to_date": ["hours to date", "hours_to_date", "htd", "prior hours",
                          "cumulative hours", "hours booked to date"],
        "region":           ["location", "region", "country", "office"],
        "customer_region":  ["customer region", "customer_region", "cust region", "client region"],
        "project_manager":  ["project manager", "project_manager", "pm", "manager"],
    }
    mapping = {}
    unmatched = []
    for standard, candidates in checks.items():
        for candidate in candidates:
            if candidate in cols_lower:
                mapping[cols_lower[candidate]] = standard
                break
        else:
            if standard in ["employee", "project", "project_type", "hours", "non_billable"]:
                unmatched.append(standard)
    return mapping, unmatched

# ── Core credit logic ─────────────────────────────────────────────────────────
def assign_credits(df, scope_map):
    col_map, unmatched = auto_detect_columns(df)
    if unmatched:
        st.warning(f"⚠️ Could not auto-detect columns: {unmatched}. Check your file headers.")

    df = df.rename(columns=col_map)
    df["non_billable"] = df["non_billable"].astype(str).str.strip().str.upper()
    # Normalize project names — collapse internal whitespace and strip edges
    if "project" in df.columns:
        df["project"] = df["project"].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
    df["hours"]  = pd.to_numeric(df["hours"], errors="coerce").fillna(0)
    df["date"]   = pd.to_datetime(df["date"], errors="coerce")
    df["period"] = df["date"].dt.strftime("%Y-%m").fillna("Unknown")
    # Map employee location → PS Region (APAC / EMEA / NOAM)
    if "region" in df.columns:
        df["ps_region"] = df["region"].map(PS_REGION_MAP).fillna("Other")
    else:
        df["ps_region"] = "Other"

    consumed = {}
    credit_hrs_list    = []
    variance_hrs_list  = []
    credit_tag_list    = []
    notes_list         = []
    htd_start_list     = []  # track starting HTD per row for output

    for _, row in df.iterrows():
        proj      = " ".join(str(row.get("project", "")).split())  # normalize whitespace
        ptype     = str(row.get("project_type", "")).strip()
        hrs       = float(row.get("hours", 0))
        nb        = str(row.get("non_billable", "NO")).strip().upper()
        bill_type = str(row.get("billing_type", "")).strip().lower()
        is_tm     = bill_type == "t&m"

        if hrs <= 0:
            credit_hrs_list.append(0); variance_hrs_list.append(0)
            credit_tag_list.append("SKIPPED"); notes_list.append("Zero or missing hours")
            htd_start_list.append(0)
            continue

        # Rule 1: Internal = always 0 credit
        if bill_type == "internal":
            credit_hrs_list.append(0); variance_hrs_list.append(0)
            credit_tag_list.append("NON-BILLABLE"); notes_list.append("Internal: excluded from utilization")
            htd_start_list.append(0)
            continue

        # Rule 2: T&M = always full credit, no cap
        if is_tm:
            credit_hrs_list.append(hrs); variance_hrs_list.append(0)
            credit_tag_list.append("CREDITED"); notes_list.append("T&M: full credit")
            htd_start_list.append(0)
            continue

        # Rule 3: Fixed Fee = capped at scope (longest match wins)
        _ptype_lower = ptype.strip().lower()
        _matches = [(k, float(v)) for k, v in scope_map.items()
                    if k.strip().lower() in _ptype_lower]
        scope_hrs = max(_matches, key=lambda x: len(x[0]))[1] if _matches else None

        if scope_hrs is None:
            credit_hrs_list.append(0); variance_hrs_list.append(hrs)
            credit_tag_list.append("UNCONFIGURED"); notes_list.append(f"Fixed Fee but no scope defined for: {ptype}")
            htd_start_list.append(0)
            continue

        # Seed starting balance from hours_to_date if first time seeing this project
        if proj not in consumed:
            htd = row.get("hours_to_date", None)
            try:
                consumed[proj] = float(htd) if htd is not None and str(htd).strip() not in ("", "nan") else 0
            except (ValueError, TypeError):
                consumed[proj] = 0

        already    = consumed[proj]
        remaining  = scope_hrs - already
        htd_start_list.append(already)

        if remaining <= 0:
            credit_hrs_list.append(0); variance_hrs_list.append(hrs)
            credit_tag_list.append("OVERRUN"); notes_list.append(f"Scope exhausted (cap: {scope_hrs:.0f}h)")
        elif hrs <= remaining:
            consumed[proj] = already + hrs
            credit_hrs_list.append(hrs); variance_hrs_list.append(0)
            credit_tag_list.append("CREDITED"); notes_list.append(f"NB within scope ({already:.1f}/{scope_hrs:.0f}h used)")
        else:
            consumed[proj] = already + remaining
            credit_hrs_list.append(remaining); variance_hrs_list.append(hrs - remaining)
            credit_tag_list.append("PARTIAL")
            notes_list.append(f"Split: {remaining:.2f}h credited / {hrs - remaining:.2f}h overrun")

    df["credit_hrs"]    = credit_hrs_list
    df["variance_hrs"]  = variance_hrs_list
    df["credit_tag"]    = credit_tag_list
    df["notes"]         = notes_list
    df["htd_start"]     = htd_start_list

    # Updated hours to date = htd_start + total hours booked this period (credited + overrun)
    df["previous_htd"] = df["htd_start"] + df["hours"]

    # Tag FF tasks
    df["ff_task"] = df["task"].apply(match_ff_task) if "task" in df.columns else None

    # Collect skipped rows for reporting
    skipped_df = df[df["credit_tag"] == "SKIPPED"][
        [c for c in ["employee","project","project_type","billing_type","date","hours","notes"] if c in df.columns]
    ].copy()

    return df, consumed, skipped_df


# ── Excel builder ─────────────────────────────────────────────────────────────
def build_excel(df, scope_map, consumed):
    wb  = Workbook()
    wb.remove(wb.active)
    bgs = [WHITE, LTGRAY]

    # ── 1. PROCESSED DATA ─────────────────────────────────────

    ws = wb.create_sheet("PROCESSED_DATA")
    ws.sheet_properties.tabColor = TEAL
    ws.freeze_panes = "A3"

    headers = ["Employee","Location","Customer Region","Project Manager","Project",
               "Project Type","Billing Type","Hrs to Date","Date","Hours Logged",
               "Approval","Task/Case","Non-Billable","Credit Hrs","Variance Hrs",
               "Previous Hrs to Date","Credit Tag","Period","Notes"]
    widths  = [20,16,18,20,35,20,14,13,14,13,14,25,13,12,12,18,16,12,45]
    cols    = ["employee","region","customer_region","project_manager","project",
               "project_type","billing_type","hours_to_date","date","hours",
               "approval","task","non_billable","credit_hrs","variance_hrs",
               "previous_htd","credit_tag","period","notes"]

    write_title(ws, "PROCESSED DATA — Utilization Credit Detail", len(headers))
    style_header(ws, 2, headers, TEAL)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        tag = str(row.get("credit_tag","")).strip()
        bg  = TAG_COLORS.get(tag, "F2F2F2")
        for c_idx, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            fmt, bold, align = None, False, "left"
            if col == "date" and pd.notna(val):
                fmt = "YYYY-MM-DD"; align = "center"
            elif col in ("hours","credit_hrs","variance_hrs","hours_to_date","previous_htd"):
                fmt = "#,##0.00"; align = "right"
            elif col == "credit_tag":
                bold = True; align = "center"
            elif col in ("period","billing_type","region"):
                align = "center"
            style_cell(cell, bg, fmt=fmt, bold=bold, align=align)

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # ── 2. EMPLOYEE SUMMARY ───────────────────────────────────
    ws2 = wb.create_sheet("SUMMARY - By Employee")
    ws2.sheet_properties.tabColor = NAVY
    ws2.freeze_panes = "A3"

    eh = ["Employee","Location","Period",
          "Avail Hrs","Hours This Period","Utilization Credits","FF Project Overrun Hrs","Admin Hrs","Util %"]
    ew = [22,16,12,12,15,18,18,14,10]
    write_title(ws2, "SUMMARY — Utilization by Employee", len(eh))
    style_header(ws2, 2, eh, TEAL)
    ws2.auto_filter.ref = "A2:I2"

    for i, w in enumerate(ew, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Get region per employee (first occurrence)
    emp_region = {}
    emp_cust_region = {}
    emp_pm = {}
    if "region" in df.columns:
        emp_region = df.dropna(subset=["region"]).groupby("employee")["region"].first().to_dict()
    if "customer_region" in df.columns:
        emp_cust_region = df.dropna(subset=["customer_region"]).groupby("employee")["customer_region"].first().to_dict()
    if "project_manager" in df.columns:
        emp_pm = df.dropna(subset=["project_manager"]).groupby("employee")["project_manager"].first().to_dict()

    # Admin hours = all Internal billing type rows
    admin_hrs = df[df["billing_type"].str.lower() == "internal"].groupby(
        ["employee","period"])["hours"].sum().reset_index().rename(columns={"hours":"admin_hrs"})

    emp_sum = df[df["credit_tag"] != "SKIPPED"].groupby(
        ["employee","period"], as_index=False
    ).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        ff_overrun_hrs=("variance_hrs","sum"),
    ).sort_values(["employee","period"])

    emp_sum = emp_sum.merge(admin_hrs, on=["employee","period"], how="left")
    emp_sum["admin_hrs"] = emp_sum["admin_hrs"].fillna(0)

    _prev_emp = None; _grp_idx = 0
    for r_idx, (_, row) in enumerate(emp_sum.iterrows(), 3):
        emp     = row["employee"]
        period  = row["period"]
        region  = emp_region.get(emp, "")
        avail   = get_avail_hours(region, period) if region else None
        util    = row["credit_hrs"] / row["hours_this_period"] if row["hours_this_period"] > 0 else 0
        bg, _grp_idx = group_bg(emp, _prev_emp, _grp_idx)
        _prev_emp = emp
        util_bg = ("EAF9F1" if util >= 0.8 else "FEF9E7" if util >= 0.6 else "FDECED")

        vals = [emp, region, period, avail or "—",
                row["hours_this_period"], row["credit_hrs"],
                row["ff_overrun_hrs"], row.get("admin_hrs", 0),
                util if avail else "—"]
        fmts = [None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%"]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, util_bg if c_idx == 9 else bg, fmt=fmt,
                       align="right" if c_idx > 3 else "center" if c_idx == 3 else "left")

    # ── 3. PROJECT SUMMARY ────────────────────────────────────
    ws3 = wb.create_sheet("SUMMARY - By Project")
    ws3.sheet_properties.tabColor = "E67E22"
    ws3.freeze_panes = "A3"

    ph = ["Project","Project Type","Customer Region","Project Manager",
          "Scoped Hrs","Hours to Date","Hours This Period","Credit Hrs",
          "FF Project Overrun Hrs","Previous Hrs to Date","Burn %","Status"]
    pw = [35,20,18,20,12,15,15,12,18,18,10,12]
    write_title(ws3, "SUMMARY — Utilization by Project", len(ph))
    style_header(ws3, 2, ph, TEAL)
    ws3.auto_filter.ref = "A2:K2"

    for i, w in enumerate(pw, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Project Summary: Fixed Fee only
    ff_proj_df = df[
        (df["credit_tag"] != "SKIPPED") &
        (df["billing_type"].str.lower() == "fixed fee")
    ] if "billing_type" in df.columns else df[df["credit_tag"] != "SKIPPED"]

    proj_sum = ff_proj_df.groupby(
        ["project","project_type"], as_index=False
    ).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        variance_hrs=("variance_hrs","sum"),
        htd_start=("htd_start","first"),
    ).sort_values("project")

    # HTD seed now comes directly from aggregation above
    htd_seeds = dict(zip(proj_sum["project"], proj_sum["htd_start"]))

    # Project-level lookups for Customer Region and Project Manager
    proj_cust_region = {}
    proj_pm = {}
    if "customer_region" in df.columns:
        proj_cust_region = df.dropna(subset=["customer_region"]).groupby("project")["customer_region"].first().to_dict()
    if "project_manager" in df.columns:
        proj_pm = df.dropna(subset=["project_manager"]).groupby("project")["project_manager"].first().to_dict()

    _prev_ptype = None; _grp_idx_p = 0
    for r_idx, (_, row) in enumerate(proj_sum.iterrows(), 3):
        ptype   = str(row["project_type"]).strip()
        _pm     = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in ptype.lower()]
        scope_h = max(_pm, key=lambda x: len(x[0]))[1] if _pm else 0

        seed      = float(row["htd_start"]) if row["htd_start"] else 0
        # htd_start already includes hours this period (per NetSuite export)
        # Previous HTD = htd_start minus hours booked this period
        previous_h = seed - row["hours_this_period"]

        # Burn % = htd_start / scoped hrs (htd_start includes this period)
        burn = seed / scope_h if scope_h > 0 else 0

        vari_h  = row["variance_hrs"]
        if vari_h > 0 or burn > 1:  status = "OVERRUN"
        elif burn >= 0.9:            status = "REVIEW"
        elif burn > 0:               status = "ON TRACK"
        else:                        status = "—"

        status_bg = {"OVERRUN":"FDECED","REVIEW":"FEF9E7","ON TRACK":"EAF9F1"}.get(status, LTGRAY)
        bg, _grp_idx_p = group_bg(ptype, _prev_ptype, _grp_idx_p)
        _prev_ptype = ptype

        cust_reg = proj_cust_region.get(row["project"], "")
        pm_name  = proj_pm.get(row["project"], "")
        vals = [row["project"], ptype, cust_reg, pm_name, scope_h or "—", previous_h,
                row["hours_this_period"], row["credit_hrs"], vari_h, seed,
                burn if scope_h > 0 else "—", status]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%",None]

        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, status_bg if c_idx == 12 else bg,
                       fmt=fmt, bold=(c_idx == 12),
                       align="right" if c_idx in (5,6,7,8,9,10,11) else "center" if c_idx == 12 else "left")

    # ── 4. ZCO NON-BILLABLE BREAKDOWN ─────────────────────────
    ws5 = wb.create_sheet("ZCO Non-Billable")
    ws5.sheet_properties.tabColor = "95A5A6"
    ws5.freeze_panes = "A3"

    znh = ["Task / Activity","Employee","Period","Hours"]
    znw = [35,22,12,12]
    write_title(ws5, "ZCO NON-BILLABLE — Hours by Employee by Activity", len(znh))
    style_header(ws5, 2, znh, TEAL)
    ws5.auto_filter.ref = "A2:E2"

    for i, w in enumerate(znw, 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    zco_df = df[df["credit_tag"] == "NON-BILLABLE"].copy()
    if "task" in zco_df.columns and len(zco_df) > 0:
        zco_sum = zco_df.groupby(
            ["task","employee","period"], as_index=False
        ).agg(hours=("hours","sum")).sort_values(["task","employee","period"])

        # Total hours per employee per period (all rows incl billable)
        emp_period_totals = df[df["credit_tag"] != "SKIPPED"].groupby(
            ["employee","period"])["hours"].sum().to_dict()

        znh[3] = "Hours"
        # Update headers to include % col
        ws5.cell(row=2, column=5, value="% of Total Hrs").font = Font(name="Manrope", bold=True, color=WHITE, size=10)
        ws5.cell(row=2, column=5).fill = hdr_fill(TEAL)
        ws5.cell(row=2, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws5.cell(row=2, column=5).border = thin_border()
        ws5.column_dimensions["E"].width = 16

        _prev_task_z = None; _grp_idx_z = 0; r_idx = 3
        for _, row in zco_sum.iterrows():
            task = row.get("task","")
            # Navy section header when task changes
            if task != _prev_task_z:
                _task_hrs = zco_sum[zco_sum["task"]==task]["hours"].sum()
                for ci, (hval, hfmt) in enumerate([
                    (task, None), ("", None), ("", None),
                    (_task_hrs, "#,##0.00"), ("", None)], 1):
                    hcell = ws5.cell(row=r_idx, column=ci, value=hval)
                    hcell.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
                    hcell.fill  = PatternFill("solid", fgColor=NAVY if ci==1 else "D6DCF0")
                    hcell.border = thin_border()
                    hcell.alignment = Alignment(
                        horizontal="right" if ci==4 else "left", vertical="center")
                    if hfmt: hcell.number_format = hfmt
                r_idx += 1
                _prev_task_z = task
                _grp_idx_z = 0
            bg, _grp_idx_z = group_bg(task, task, _grp_idx_z)
            total_hrs  = emp_period_totals.get((row["employee"], row["period"]), 0)
            pct        = row["hours"] / total_hrs if total_hrs > 0 else 0
            vals = ["", row["employee"], row["period"], row["hours"], pct]
            fmts = [None, None, None, "#,##0.00", "0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=val)
                style_cell(cell, bg, fmt=fmt,
                           align="right" if c_idx in (4,5) else "center" if c_idx == 3 else "left")
            r_idx += 1
    else:
        ws5.cell(row=3, column=1, value="No Non-Billable (Internal) entries in this period.")

    # ── 5. TASK ANALYSIS ───────────────────────
    ws6 = wb.create_sheet("Task Analysis")
    ws6.sheet_properties.tabColor = "27AE60"
    ws6.freeze_panes = "A3"

    tah = ["Task Category","Project Type","Hours This Period","Avg Hrs / Project","% of Type Hrs"]
    taw = [28,25,16,18,16]
    write_title(ws6, "TASK ANALYSIS — Hours by Task › Project Type", len(tah))
    style_header(ws6, 2, tah, TEAL)
    ws6.auto_filter.ref = "A2:E2"

    for i, w in enumerate(taw, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    # Only Fixed Fee rows with a matched FF task
    ff_df = df[(df["billing_type"].str.lower() == "fixed fee") & (df["ff_task"].notna())].copy() \
        if "billing_type" in df.columns else df[df["ff_task"].notna()].copy()

    if len(ff_df) > 0:
        task_sum = ff_df.groupby(
            ["ff_task","project_type"], as_index=False
        ).agg(
            hours=("hours","sum"),
        ).sort_values(["ff_task","project_type"])

        # Distinct project count per type using ALL fixed fee rows (not just task-tagged rows)
        all_ff = df[df["billing_type"].str.lower() == "fixed fee"] if "billing_type" in df.columns else df
        proj_count_by_type = all_ff.groupby("project_type")["project"].nunique().to_dict()

        # Total hours per type for % calc
        type_totals = ff_df.groupby("project_type")["hours"].sum().to_dict()

        _prev_task_t = None; _grp_idx_t = 0; r_idx_t = 3
        for _, row in task_sum.iterrows():
            ff_task = row["ff_task"]
            type_total = type_totals.get(row["project_type"], 0)
            pct        = row["hours"] / type_total if type_total > 0 else 0
            proj_cnt   = proj_count_by_type.get(row["project_type"], 1)
            raw_avg    = row["hours"] / proj_cnt if proj_cnt > 0 else 0
            avg_hrs    = round(raw_avg * 4) / 4
            # Navy section header when task changes
            if ff_task != _prev_task_t:
                _task_total_hrs = task_sum[task_sum["ff_task"]==ff_task]["hours"].sum()
                for ci, (hval, hfmt) in enumerate([
                    (ff_task, None), ("— ALL TYPES —", None),
                    (_task_total_hrs, "#,##0.00"), ("", None), ("", None)], 1):
                    hcell = ws6.cell(row=r_idx_t, column=ci, value=hval)
                    hcell.font  = Font(name="Manrope", size=10, bold=True, color="FFFFFF")
                    hcell.fill  = PatternFill("solid", fgColor=NAVY if ci<=2 else "D6DCF0")
                    hcell.border = thin_border()
                    hcell.alignment = Alignment(
                        horizontal="right" if ci==3 else "left", vertical="center")
                    if hfmt: hcell.number_format = hfmt
                r_idx_t += 1
                _prev_task_t = ff_task
                _grp_idx_t = 0
            task_colors = {
                "Configuration":        "EBF5FB",
                "Enablement/Training":  "EAF9F1",
                "Post Go-live Support": "FEF9E7",
                "Project Management":   "F4ECF7",
            }
            bg, _grp_idx_t = group_bg(ff_task, ff_task, _grp_idx_t)
            task_bg = task_colors.get(ff_task, bg)
            vals = ["", row["project_type"], row["hours"], avg_hrs, pct]
            fmts = [None, None, "#,##0.00", "#,##0.00", "0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws6.cell(row=r_idx_t, column=c_idx, value=val)
                style_cell(cell, task_bg if c_idx > 1 else bg, fmt=fmt,
                           align="right" if c_idx > 2 else "left")
            r_idx_t += 1
    else:
        ws6.cell(row=3, column=1, value="No Fixed Fee task data found. Check Billing Type and Task/Case columns.")

    # ── 6. PROJECT COUNT BY TYPE ─────────────────────────────
    ws_pc = wb.create_sheet("Project Count")
    ws_pc.sheet_properties.tabColor = "2980B9"
    ws_pc.freeze_panes = "A3"

    pch = ["Project Type","Billing Type","Project Count"]
    pcw = [35, 14, 14]
    write_title(ws_pc, "PROJECT COUNT — Distinct Projects by Type (excl. Internal)", len(pch))
    style_header(ws_pc, 2, pch, TEAL)
    ws_pc.auto_filter.ref = "A2:D2"

    for i, w in enumerate(pcw, 1):
        ws_pc.column_dimensions[get_column_letter(i)].width = w

    # Exclude internal, count distinct projects per type only
    pc_df = df[df["billing_type"].str.lower() != "internal"].copy()         if "billing_type" in df.columns else df.copy()

    pc_sum = pc_df.groupby(["project_type","billing_type"], as_index=False).agg(
        project_count=("project","nunique"),
    ).sort_values(["project_type","billing_type"])

    grand_total = pc_sum["project_count"].sum()

    for r_idx, (_, row) in enumerate(pc_sum.iterrows(), 3):
        bg = LTGRAY if r_idx % 2 == 0 else WHITE
        vals = [row["project_type"], row["billing_type"], row["project_count"]]
        fmts = [None, None, "#,##0"]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_pc.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, bg, fmt=fmt, align="center" if c_idx == 2 else "right" if c_idx == 3 else "left")

    # Grand total row
    total_row = r_idx + 1 if len(pc_sum) > 0 else 3
    for c_idx, (val, fmt, bold) in enumerate([
        ("Grand Total", None, True),
        ("", None, False),
        (grand_total, "#,##0", True),
    ], 1):
        cell = ws_pc.cell(row=total_row, column=c_idx, value=val)
        cell.font   = Font(name="Manrope", bold=True, size=10, color=WHITE)
        cell.fill   = hdr_fill(NAVY)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="right" if c_idx == 2 else "left", vertical="center")
        if fmt:
            cell.number_format = fmt

    # ── 7. CUSTOMER REGION SUMMARY ───────────────────────────
    ws_cr = wb.create_sheet("By Customer Region")
    ws_cr.sheet_properties.tabColor = "1e2c63"
    ws_cr.freeze_panes = "A3"

    crh = ["Customer Region","Hours This Period","Utilization Credits",
           "FF Project Overrun Hrs","Util %"]
    crw = [22,16,18,20,10]
    write_title(ws_cr, "SUMMARY — Utilization by Customer Region", len(crh))
    style_header(ws_cr, 2, crh, TEAL)
    ws_cr.auto_filter.ref = "A2:E2"

    for i, w in enumerate(crw, 1):
        ws_cr.column_dimensions[get_column_letter(i)].width = w

    if "customer_region" in df.columns:
        cr_base = df[df["credit_tag"] != "SKIPPED"].copy()
        cr_base["customer_region"] = cr_base["customer_region"].fillna("Unassigned")

        cr_sum = cr_base.groupby("customer_region", as_index=False).agg(
            hours_this_period=("hours","sum"),
            credit_hrs=("credit_hrs","sum"),
            ff_overrun_hrs=("variance_hrs","sum"),
        ).sort_values("customer_region")

        for r_idx, (_, row) in enumerate(cr_sum.iterrows(), 3):
            cr      = row["customer_region"]
            total_h = row["hours_this_period"]
            util    = row["credit_hrs"] / total_h if total_h > 0 else 0
            util_bg = ("EAF9F1" if util >= 0.8 else "FEF9E7" if util >= 0.6 else "FDECED")
            bg      = bgs[r_idx % 2]
            vals = [cr, total_h, row["credit_hrs"], row["ff_overrun_hrs"],
                    util if total_h > 0 else "—"]
            fmts = [None,"#,##0.00","#,##0.00","#,##0.00","0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
                cell = ws_cr.cell(row=r_idx, column=c_idx, value=val)
                style_cell(cell, util_bg if c_idx == 5 else bg, fmt=fmt,
                           align="right" if c_idx > 1 else "left")
    else:
        ws_cr.cell(row=3, column=1, value="No 'Customer Region' column found in import.")

    # ── 8. PS REGION SUMMARY ─────────────────────────────────
    ws_ps = wb.create_sheet("By PS Region")
    ws_ps.sheet_properties.tabColor = "4472C4"
    ws_ps.freeze_panes = "A4"

    psh = ["PS Region","Project Type","Billing Type",
           "Avail Hrs","Hours This Period","Utilization Credits",
           "FF Project Overrun Hrs","Admin Hrs","Util %"]
    psw = [14,28,14,12,16,18,20,14,10]
    write_title(ws_ps, "SUMMARY — Utilization by PS Region (APAC / EMEA / NOAM)", len(psh))
    style_header(ws_ps, 2, psh, TEAL)
    ws_ps.auto_filter.ref = "A3:I3"

    # Sub-header note
    ws_ps.cell(row=3, column=1,
        value="Grouped by PS Region → Project Type → Billing Type").font = Font(
        name="Manrope", size=9, italic=True, color="808080")
    for i, w in enumerate(psw, 1):
        ws_ps.column_dimensions[get_column_letter(i)].width = w

    # Avail hrs per region (unique employee+period)
    ps_avail = {}
    _seen_ep = set()
    for _emp, _grp in df.groupby("employee"):
        _loc  = emp_region.get(_emp, "")
        _ps   = PS_REGION_MAP.get(_loc, "Other")
        for _p in _grp["period"].unique():
            if (_emp, _p) not in _seen_ep:
                _seen_ep.add((_emp, _p))
                ps_avail[_ps] = ps_avail.get(_ps, 0) + (get_avail_hours(_loc, _p) or 0)

    # Admin hrs per region
    ps_admin = {}
    if "billing_type" in df.columns:
        for _, _ar in df[df["billing_type"].str.lower()=="internal"].iterrows():
            _ps = PS_REGION_MAP.get(_ar.get("region",""), "Other")
            ps_admin[_ps] = ps_admin.get(_ps, 0) + _ar.get("hours", 0)

    # Build 3-level aggregation
    _ps_base = df[df["credit_tag"] != "SKIPPED"].copy()
    if "billing_type" not in _ps_base.columns:
        _ps_base["billing_type"] = "Unknown"
    _ps_detail = _ps_base.groupby(
        ["ps_region","project_type","billing_type"], as_index=False
    ).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        ff_overrun_hrs=("variance_hrs","sum"),
    )

    # Region subtotals
    _ps_reg_total = _ps_base.groupby("ps_region", as_index=False).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        ff_overrun_hrs=("variance_hrs","sum"),
    )

    region_order = ["APAC","EMEA","NOAM","Other"]
    _ps_detail["_rord"] = _ps_detail["ps_region"].map(
        {r:i for i,r in enumerate(region_order)}).fillna(99)
    _ps_detail = _ps_detail.sort_values(
        ["_rord","ps_region","project_type","billing_type"]).drop(columns=["_rord"])

    r_idx = 4
    _last_region = None
    for _, row in _ps_detail.iterrows():
        ps_reg = row["ps_region"]

        # ── Region header row ──────────────────────────────────
        if ps_reg != _last_region:
            _last_region = ps_reg
            _rt = _ps_reg_total[_ps_reg_total["ps_region"]==ps_reg]
            _rh = _rt.iloc[0]["hours_this_period"] if len(_rt) else 0
            _rc = _rt.iloc[0]["credit_hrs"]         if len(_rt) else 0
            _ro = _rt.iloc[0]["ff_overrun_hrs"]     if len(_rt) else 0
            _ra = ps_admin.get(ps_reg, 0)
            _rv = ps_avail.get(ps_reg, 0)
            _ru = _rc / _rh if _rh > 0 else 0
            _ru_bg = "EAF9F1" if _ru>=0.7 else "FEF9E7" if _ru>=0.6 else "FDECED"

            reg_vals = [ps_reg, "— ALL TYPES —", "",
                        _rv or "—", _rh, _rc, _ro, _ra,
                        _ru if _rh > 0 else "—"]
            reg_fmts = [None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.0%"]
            for c_idx, (val, fmt) in enumerate(zip(reg_vals, reg_fmts), 1):
                cell = ws_ps.cell(row=r_idx, column=c_idx, value=val)
                cell.font  = Font(name="Manrope", size=10, bold=True,
                                  color="FFFFFF" if c_idx <= 2 else "000000")
                cell.fill  = PatternFill("solid", fgColor=NAVY if c_idx <= 2 else (
                                  _ru_bg if c_idx == 9 else "D6DCF0"))
                cell.border = thin_border()
                if fmt: cell.number_format = fmt
                cell.alignment = Alignment(horizontal="right" if c_idx > 3 else "left",
                                           vertical="center")
            r_idx += 1

        # ── Detail row ─────────────────────────────────────────
        hrs  = row["hours_this_period"]
        util = row["credit_hrs"] / hrs if hrs > 0 else 0
        util_bg = "EAF9F1" if util >= 0.7 else "FEF9E7" if util >= 0.6 else "FDECED"
        bg = bgs[r_idx % 2]

        vals = ["", row["project_type"], row["billing_type"],
                "", hrs, row["credit_hrs"], row["ff_overrun_hrs"], "",
                util if hrs > 0 else "—"]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","#,##0.00",None,"0.0%"]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_ps.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, util_bg if c_idx == 9 else bg, fmt=fmt,
                       align="right" if c_idx > 3 else "left")
        r_idx += 1


    # ── 9. PROJECT WATCH LIST ────────────────────────────────
    ws_wl = wb.create_sheet("Watch List")
    ws_wl.sheet_properties.tabColor = "E74C3C"
    ws_wl.freeze_panes = "A3"

    # Section A: Top 10 overrun projects
    wlh = ["Project","Project Type","Customer Region","Project Manager",
           "Scoped Hrs","Previous Hrs to Date","Burn %","FF Overrun Hrs","Status"]
    wlw = [35,20,18,20,12,18,10,14,12]
    write_title(ws_wl, "PROJECT WATCH LIST — Overrun & At-Risk Projects", len(wlh))
    style_header(ws_wl, 2, wlh, "E74C3C")
    for i, w in enumerate(wlw, 1):
        ws_wl.column_dimensions[get_column_letter(i)].width = w

    # Build project-level data from ff_proj_df
    wl_df = ff_proj_df.groupby(["project","project_type"], as_index=False).agg(
        hours_this_period=("hours","sum"),
        credit_hrs=("credit_hrs","sum"),
        variance_hrs=("variance_hrs","sum"),
        htd_start=("htd_start","first"),
    )
    wl_df["previous_htd"] = wl_df.apply(
        lambda r: max(0, (float(r["htd_start"]) if r["htd_start"] else 0) - r["hours_this_period"]), axis=1)

    def get_scope(ptype):
        _pm = [(k, float(v)) for k, v in scope_map.items() if k.strip().lower() in str(ptype).strip().lower()]
        return max(_pm, key=lambda x: len(x[0]))[1] if _pm else 0

    wl_df["scope_h"]  = wl_df["project_type"].apply(get_scope)
    wl_df["burn_pct"] = wl_df.apply(
        lambda r: (float(r["htd_start"]) if r["htd_start"] else 0) / r["scope_h"] if r["scope_h"] > 0 else None, axis=1)
    wl_df["status"] = wl_df.apply(
        lambda r: "OVERRUN" if (r["variance_hrs"] > 0 or (r["burn_pct"] or 0) > 1)
        else "REVIEW" if (r["burn_pct"] or 0) >= 0.9
        else "ON TRACK", axis=1)

    # Filter to OVERRUN + AT RISK, sort by burn desc
    watchlist = wl_df[wl_df["status"].isin(["OVERRUN","REVIEW"])].sort_values(
        "burn_pct", ascending=False).head(25)

    r_idx = 3
    for _, row in watchlist.iterrows():
        status   = row["status"]
        status_bg = "FDECED" if status == "OVERRUN" else "FEF9E7"
        bg       = status_bg
        burn_val = row["burn_pct"] if row["burn_pct"] is not None else "—"
        cust_reg = proj_cust_region.get(row["project"], "")
        pm_name  = proj_pm.get(row["project"], "")
        vals = [row["project"], row["project_type"], cust_reg, pm_name,
                row["scope_h"] or "—", row["previous_htd"],
                burn_val, row["variance_hrs"], status]
        fmts = [None,None,None,None,"#,##0.00","#,##0.00","0.0%","#,##0.00",None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_wl.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, status_bg if c_idx == 9 else bg, fmt=fmt,
                       bold=(c_idx == 9),
                       align="right" if c_idx in (5,6,7,8) else "center" if c_idx == 9 else "left")
        r_idx += 1

    # Section B: FF: NO SCOPE DEFINED projects (no scope defined)
    r_idx += 1
    unconf_title_cell = ws_wl.cell(row=r_idx, column=1,
        value="FF: NO SCOPE DEFINED (Hours at Risk)")
    unconf_title_cell.font  = Font(name="Manrope", bold=True, size=11, color="FFFFFF")
    unconf_title_cell.fill  = hdr_fill("E67E22")
    ws_wl.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(wlh))
    r_idx += 1

    unconf_df = df[df["credit_tag"] == "UNCONFIGURED"].groupby(
        ["project","project_type"], as_index=False
    ).agg(hours=("hours","sum")).sort_values("hours", ascending=False)

    for _, row in unconf_df.iterrows():
        bg = "FEF3E2"
        vals = [row["project"], row["project_type"], proj_cust_region.get(row["project"],""),
                proj_pm.get(row["project"],""), "—", "—", "—", row["hours"], "FF: NO SCOPE DEFINED"]
        fmts = [None,None,None,None,None,None,None,"#,##0.00",None]
        for c_idx, (val, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws_wl.cell(row=r_idx, column=c_idx, value=val)
            style_cell(cell, bg, fmt=fmt,
                       align="right" if c_idx == 8 else "left")
        r_idx += 1

    # ── 10. SKIPPED ROWS ──────────────────────────────────────
    ws7 = wb.create_sheet("Skipped Rows")
    ws7.sheet_properties.tabColor = "E74C3C"
    ws7.freeze_panes = "A3"

    skh = ["Employee","Project","Project Type","Billing Type","Date","Hours","Reason"]
    skw = [20,35,20,14,14,10,45]
    write_title(ws7, "SKIPPED ROWS — Not Included in Utilization Calculations", len(skh))
    style_header(ws7, 2, skh, "C0392B")
    for i, w in enumerate(skw, 1):
        ws7.column_dimensions[get_column_letter(i)].width = w

    skip_cols = ["employee","project","project_type","billing_type","date","hours","notes"]
    skipped = df[df["credit_tag"] == "SKIPPED"]
    if len(skipped) > 0:
        for r_idx, (_, row) in enumerate(skipped.iterrows(), 3):
            for c_idx, col in enumerate(skip_cols, 1):
                val  = row.get(col, "")
                cell = ws7.cell(row=r_idx, column=c_idx, value=val)
                fmt  = "YYYY-MM-DD" if col == "date" else "#,##0.00" if col == "hours" else None
                style_cell(cell, "FDECED", fmt=fmt,
                           align="right" if col == "hours" else "center" if col == "date" else "left")
    else:
        ws7.cell(row=3, column=1, value="No skipped rows — all entries were processed.")

    # ── DASHBOARD (exec summary) ─────────────────────────────
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.sheet_properties.tabColor = "1e2c63"
    ws_dash.sheet_view.showGridLines = False

    def dash_label(ws, row, col, text, size=10, bold=False, color="808080"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        return c

    def dash_value(ws, row, col, value, fmt=None, size=18, bold=True, color="1e2c63"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Manrope", size=size, bold=bold, color=color)
        if fmt: c.number_format = fmt
        return c

    def dash_section(ws, row, col, text, ncols=4):
        c = ws.cell(row=row, column=col, value=text)
        c.font  = Font(name="Manrope", size=11, bold=True, color="FFFFFF")
        c.fill  = hdr_fill(NAVY)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ncols-1)
        return c

    def rag_cell(ws, row, col, value, fmt=None, status="green"):
        colors = {"green":"EAF9F1","yellow":"FEF9E7","red":"FDECED"}
        txt    = {"green":"2ECC71","yellow":"F39C12","red":"E74C3C"}
        c = ws.cell(row=row, column=col, value=value)
        c.font  = Font(name="Manrope", size=14, bold=True, color=txt.get(status,"000000"))
        c.fill  = PatternFill("solid", fgColor=colors.get(status,"FFFFFF"))
        c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt: c.number_format = fmt
        return c

    for col, w in [(1,3),(2,22),(3,18),(4,18),(5,18),(6,18),(7,18),(8,3)]:
        ws_dash.column_dimensions[get_column_letter(col)].width = w
    for row in range(1, 45):
        ws_dash.row_dimensions[row].height = 18

    # Title
    tc = ws_dash.cell(row=2, column=2, value="Professional Services — Utilization Credit Report")
    tc.font = Font(name="Manrope", size=16, bold=True, color="FFFFFF")
    tc.fill = hdr_fill(NAVY)
    ws_dash.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws_dash.row_dimensions[2].height = 30

    if "date" in df.columns:
        max_dt   = pd.to_datetime(df["date"], errors="coerce").max()
        date_str = max_dt.strftime("%d %B %Y") if pd.notna(max_dt) else "—"
    else:
        date_str = "—"
    sc = ws_dash.cell(row=3, column=2, value=f"Data through {date_str}")
    sc.font = Font(name="Manrope", size=10, color="808080")
    ws_dash.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    kc = ws_dash.cell(row=4, column=2, value="This report calculates Utilization Credits from NetSuite time detail exports. T&M projects: full credit for all hours logged. Fixed Fee projects: credit up to scoped hours; hours beyond scope tracked as overrun (excluded from credits). Internal time: excluded from utilization, tracked as Admin Hours. Util % = Utilization Credits / Hours This Period.")
    kc.font = Font(name="Manrope", size=9, italic=True, color="808080")
    kc.alignment = Alignment(wrap_text=True)
    ws_dash.merge_cells(start_row=4, start_column=2, end_row=4, end_column=7)
    ws_dash.row_dimensions[4].height = 30
    # Push all section starts down 1 row to accommodate key text


    # Key Metrics
    dash_section(ws_dash, 6, 2, "KEY METRICS", ncols=6)
    ws_dash.row_dimensions[5].height = 22
    hours_tp_d    = df[df["credit_tag"] != "SKIPPED"]["hours"].sum()
    credit_hrs_d  = df[df["credit_tag"].isin(["CREDITED","PARTIAL"])]["credit_hrs"].sum()
    overrun_hrs_d = df[df["credit_tag"] == "OVERRUN"]["variance_hrs"].sum()
    admin_hrs_d   = df[df["billing_type"].str.lower()=="internal"]["hours"].sum() if "billing_type" in df.columns else 0
    total_rows_d  = len(df[df["credit_tag"] != "SKIPPED"])
    util_pct_d    = credit_hrs_d / hours_tp_d if hours_tp_d > 0 else 0
    util_status_d = "green" if util_pct_d >= 0.70 else "yellow" if util_pct_d >= 0.60 else "red"

    for i, (label, value, fmt, status) in enumerate([
        ("Hours This Period", hours_tp_d, "#,##0.00", None),
        ("Utilization Credits", credit_hrs_d, "#,##0.00", None),
        ("Util % (target 70%)", util_pct_d, "0.0%", util_status_d),
        ("FF Overrun Hrs", overrun_hrs_d, "#,##0.00", None),
        ("Admin Hrs", admin_hrs_d, "#,##0.00", None),
        ("Rows Processed", total_rows_d, "#,##0", None),
    ]):
        col = 2 + i
        dash_label(ws_dash, 7, col, label)
        if status:
            rag_cell(ws_dash, 8, col, value, fmt=fmt, status=status)
        else:
            dash_value(ws_dash, 8, col, value, fmt=fmt, size=14)
    ws_dash.row_dimensions[8].height = 28

    # PS Region
    dash_section(ws_dash, 10, 2, "UTILIZATION BY PS REGION", ncols=6)
    ws_dash.row_dimensions[9].height = 22
    for ci, hdr in enumerate(["PS Region","Hours This Period","Credit Hrs","Util %","FF Overrun Hrs","Admin Hrs"], 2):
        c = ws_dash.cell(row=11, column=ci, value=hdr)
        c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
        c.fill = hdr_fill(TEAL)

    ps_base_d = df[df["credit_tag"] != "SKIPPED"]
    ps_sum_d  = ps_base_d.groupby("ps_region").agg(
        hours=("hours","sum"), credit=("credit_hrs","sum"), overrun=("variance_hrs","sum"))
    ps_admin_d = df[df["billing_type"].str.lower()=="internal"].groupby("ps_region")["hours"].sum() if "billing_type" in df.columns else pd.Series(dtype=float)
    ps_avail_d = {}
    _seen_emp_period = set()
    for _emp2, _grp2 in df.groupby("employee"):
        _loc2  = emp_region.get(_emp2,"")
        _ps2   = PS_REGION_MAP.get(_loc2,"Other")
        for _p2 in _grp2["period"].unique():
            if (_emp2, _p2) not in _seen_emp_period:
                _seen_emp_period.add((_emp2, _p2))
                ps_avail_d[_ps2] = ps_avail_d.get(_ps2,0) + (get_avail_hours(_loc2,_p2) or 0)

    for ri, reg in enumerate(["APAC","EMEA","NOAM","Other"], 12):
        if reg not in ps_sum_d.index: continue
        _row = ps_sum_d.loc[reg]
        _adm = float(ps_admin_d.get(reg,0)) if reg in ps_admin_d.index else 0
        _avl = ps_avail_d.get(reg,0)
        _util= _row["credit"] / _row["hours"] if _row["hours"] > 0 else None
        _us  = "green" if _util>=0.70 else "yellow" if _util>=0.60 else "red"
        _bg  = bgs[ri % 2]
        _util_color = ("E74C3C" if _util<0.60 else "2ECC71" if _util>=0.70 else "F39C12") if _util is not None else "808080"
        _dash_ps_vals = [
            (2, reg,                    None,        False, "000000"),
            (3, _row["hours"],          "#,##0.00",  False, "000000"),
            (4, _row["credit"],         "#,##0.00",  False, "000000"),
            (5, _util if _util is not None else "—", "0.0%" if _util is not None else None, True, _util_color),
            (6, _row["overrun"],        "#,##0.00",  False, "000000"),
            (7, _adm,                   "#,##0.00",  False, "000000"),
        ]
        for ci2, val2, fmt2, bold2, color2 in _dash_ps_vals:
            _c = ws_dash.cell(row=ri, column=ci2, value=val2)
            _c.font  = Font(name="Manrope", size=10, bold=bold2, color=color2)
            _c.fill  = PatternFill("solid", fgColor=_bg)
            _c.border = thin_border()
            if fmt2: _c.number_format = fmt2

    # Watch List summary
    dash_section(ws_dash, 17, 2, "WATCH LIST SUMMARY", ncols=6)
    ws_dash.row_dimensions[17].height = 22
    n_overrun  = len(df[df["credit_tag"]=="OVERRUN"]["project"].unique())
    _wl_at_risk = wl_df[(wl_df["burn_pct"].notna()) & (wl_df["burn_pct"]>=0.9) & (wl_df["status"]!="OVERRUN")] if "wl_df" in dir() else pd.DataFrame()
    n_at_risk  = len(_wl_at_risk["project"].unique()) if len(_wl_at_risk) > 0 else 0
    n_unconf   = len(df[df["credit_tag"]=="UNCONFIGURED"]["project"].unique())
    unconf_hrs_d = df[df["credit_tag"]=="UNCONFIGURED"]["hours"].sum()

    for i, (label, value, fmt, status) in enumerate([
        ("Projects in Overrun", n_overrun, "#,##0", "red" if n_overrun>0 else "green"),
        ("Projects (≥90% burn)", n_at_risk, "#,##0", "yellow" if n_at_risk>0 else "green"),
        ("FF: No Scope Defined Projects", n_unconf, "#,##0", "yellow" if n_unconf>0 else "green"),
        ("FF: No Scope Defined Hours", unconf_hrs_d, "#,##0.00", "yellow" if unconf_hrs_d>0 else "green"),
    ]):
        col = 2 + i
        dash_label(ws_dash, 18, col, label)
        rag_cell(ws_dash, 19, col, value, fmt=fmt, status=status)
    ws_dash.row_dimensions[19].height = 28

    # Low utilization employees
    dash_section(ws_dash, 21, 2, "EMPLOYEES BELOW 60% UTILIZATION — Action Required", ncols=6)
    ws_dash.row_dimensions[21].height = 22
    for ci, hdr in enumerate(["Employee","Location","PS Region","Period","Util %","Credit Hrs"], 2):
        _c = ws_dash.cell(row=22, column=ci, value=hdr)
        _c.font = Font(name="Manrope", size=9, bold=True, color="FFFFFF")
        _c.fill = hdr_fill(TEAL)

    _low_rows = []
    for _, _erow in emp_sum.iterrows():
        _emp3  = _erow["employee"]
        # Skip util-exempt employees
        if any(_emp3.lower().startswith(ex.lower()) for ex in UTIL_EXEMPT_EMPLOYEES):
            continue
        _loc3  = emp_region.get(_emp3,"")
        _ps3   = PS_REGION_MAP.get(_loc3,"Other")
        _p3    = _erow["period"]
        _avl3  = get_avail_hours(_loc3, _p3) or 0
        _util3 = _erow["credit_hrs"] / _avl3 if _avl3 > 0 else 0
        if _util3 < 0.60 and _avl3 > 0:
            _low_rows.append((_emp3, _loc3, _ps3, _p3, _util3, _erow["credit_hrs"]))

    for ri, (_e,_l,_ps,_per,_u,_c) in enumerate(sorted(_low_rows, key=lambda x:x[4])[:15], 23):
        for ci2, (val2,fmt2) in enumerate([(_e,None),(_l,None),(_ps,None),(_per,None),(_u,"0.0%"),(_c,"#,##0.00")], 2):
            _cv = ws_dash.cell(row=ri, column=ci2, value=val2)
            _cv.font  = Font(name="Manrope", size=10, color="E74C3C" if ci2==6 else "000000")
            _cv.fill  = PatternFill("solid", fgColor="FDECED")
            _cv.border = thin_border()
            if fmt2: _cv.number_format = fmt2



    # ── Reorder sheets: Project Count first, Processed Data last ────────────
    sheet_order = [
        "Dashboard",
        "Project Count",
        "SUMMARY - By Employee",
        "SUMMARY - By Project",
        "By Customer Region",
        "By PS Region",
        "Watch List",
        "ZCO Non-Billable",
        "Task Analysis",
        "Skipped Rows",
        "PROCESSED_DATA",
    ]
    # Rebuild workbook sheet order directly
    existing = [s for s in sheet_order if s in wb.sheetnames]
    # Any sheets not in our list go at the end
    remaining = [s for s in wb.sheetnames if s not in existing]
    wb._sheets = [wb[s] for s in existing + remaining]

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Streamlit UI ──────────────────────────────────────────────────────────────
def main():
    st.markdown("""
        <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;600;700&display=swap" rel="stylesheet">
        <style>
            html, body, [class*="css"] { font-family: 'Manrope', sans-serif !important; }
            h1, h2, h3, .stMarkdown, .stDataFrame, label, button { font-family: 'Manrope', sans-serif !important; }
        </style>
        <div style='background-color:#1e2c63;padding:24px 32px;border-radius:8px;margin-bottom:24px;font-family:Manrope,sans-serif'>
            <h1 style='color:white;margin:0;font-size:28px;font-family:Manrope,sans-serif'>Professional Services Utilization Credit Report</h1>
            <p style='color:#aac4d0;margin:6px 0 0 0;font-size:14px;font-family:Manrope,sans-serif'>
                Upload your NetSuite time detail export to generate a utilization credit report.
                &nbsp;|&nbsp; <a href="https://3838224.app.netsuite.com/app/common/search/searchresults.nl?searchid=66732&saverun=T&whence=" style='color:#7da9f0;font-family:Manrope,sans-serif;'>Report Link</a>
            </p>
            <p style='color:#8ab0c0;margin:8px 0 0 0;font-size:12px;font-family:Manrope,sans-serif;line-height:1.6;'>This tool calculates <b>Utilization Credits</b> from NetSuite time detail exports. Credits are awarded as follows: <b>T&amp;M</b> projects receive full credit for all hours logged. <b>Fixed Fee</b> projects receive credit up to their scoped hours — hours beyond scope are tracked as overrun and excluded from credits. <b>Internal</b> time is excluded from utilization entirely and tracked separately as Admin Hours. Util&nbsp;% = Utilization Credits &divide; Hours This Period.</p>
        </div>
    """, unsafe_allow_html=True)

    # ── Adaptive metric color CSS ────────────────────────────
    st.markdown("""<style>
    :root { --text-color: #111111; }
    @media (prefers-color-scheme: dark) { :root { --text-color: #ffffff; } }
    [data-theme="dark"] { --text-color: #ffffff; }
    [data-theme="light"] { --text-color: #111111; }
    </style>""", unsafe_allow_html=True)

    # ── Upload ────────────────────────────────────────────────
    st.subheader("Step 1 — Upload NetSuite Time Detail Export")
    st.caption("Supported columns: Employee, Region, Project, Project Type, Billing Type, "
               "Hours to Date, Date, Hours, Approval Status, Case/Task/Event, Non-Billable")

    uploaded = st.file_uploader(
        "Drop your file here or click to browse",
        type=["csv", "xlsx", "xls"],
        help="Supports CSV and Excel files exported from NetSuite"
    )

    if not uploaded:
        # Show stored config as reference
        with st.expander("📋 View stored scope map & available hours"):
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Fixed Fee Scope Hours**")
                scope_df = pd.DataFrame(list(DEFAULT_SCOPE.items()), columns=["Project Type","Scoped Hrs"])
                st.dataframe(scope_df, hide_index=True, use_container_width=True)
            with c2:
                st.markdown("**Available Hours by Region (2026)**")
                avail_df = pd.DataFrame([
                    {"Region": r, **months} for r, months in AVAIL_HOURS.items()
                ])
                st.dataframe(avail_df, hide_index=True, use_container_width=True)
        st.info("👆 Upload your NetSuite export to continue.")
        return

    # Load file
    try:
        ext = os.path.splitext(uploaded.name)[1].lower()
        df_raw = pd.read_excel(uploaded) if ext in (".xlsx",".xls") else \
                 pd.read_csv(uploaded, encoding="utf-8") if True else \
                 pd.read_csv(uploaded, encoding="latin-1")
    except Exception:
        try:
            df_raw = pd.read_csv(uploaded, encoding="latin-1")
        except Exception as e:
            st.error(f"Could not read file: {e}"); return

    st.success(f"✅ Loaded **{len(df_raw):,} rows** from `{uploaded.name}`")
    with st.expander("Preview raw data (first 5 rows)"):
        st.dataframe(df_raw.head(), use_container_width=True)

    st.divider()

    # ── Process ───────────────────────────────────────────────
    st.subheader("Step 2 — Generate Report")

    if st.button("▶️ Run Utilization Engine", type="primary"):
        with st.spinner("Processing..."):
            try:
                df, consumed, skipped_df = assign_credits(df_raw.copy(), DEFAULT_SCOPE)
            except Exception as e:
                st.error(f"Processing error: {e}"); return

        st.success("✅ Processing complete!")

        # Metrics
        total_rows     = len(df[df["credit_tag"] != "SKIPPED"])
        total_credit   = df["credit_hrs"].sum()
        total_variance = df["variance_hrs"].sum()
        total_nb       = len(df[df["credit_tag"] == "NON-BILLABLE"])
        total_overrun  = len(df[df["credit_tag"] == "OVERRUN"])
        hours_this_period = df[df["credit_tag"] != "SKIPPED"]["hours"].sum() if "hours" in df.columns else 0
        total_admin    = df[df["billing_type"].str.lower() == "internal"]["hours"].sum()             if "billing_type" in df.columns else 0
        total_proj_overrun = df[df["credit_tag"] == "OVERRUN"]["variance_hrs"].sum()             if "variance_hrs" in df.columns else 0

        credit_pct  = total_credit       / hours_this_period if hours_this_period else 0
        overrun_pct = total_proj_overrun / hours_this_period if hours_this_period else 0
        admin_pct   = total_admin        / hours_this_period if hours_this_period else 0

        credit_color = "#2ecc71" if credit_pct >= 0.70 else "#f39c12" if credit_pct >= 0.60 else "#e74c3c"
        credit_label = "On target" if credit_pct >= 0.70 else "Below target" if credit_pct >= 0.60 else "At risk"

        # Max date in report
        if "date" in df.columns:
            max_date = pd.to_datetime(df["date"], errors="coerce").max()
            date_str = max_date.strftime("%-d %B %Y") if pd.notna(max_date) else "—"
        else:
            date_str = "—"
        st.markdown(f"<div style='font-size:13px;color:#a0a0a0;font-family:Manrope,sans-serif;margin-bottom:12px'>Data through <strong style='color:#ffffff'>{date_str}</strong></div>", unsafe_allow_html=True)

        m1,m2,m3,m4,m5 = st.columns(5)
        def metric_card(label, value, pill_txt=None, pill_fg=None):
            pill = ""
            if pill_txt and pill_fg:
                pill = f"<div style='display:inline-block;margin-top:6px;padding:2px 10px;border-radius:999px;background-color:{pill_fg}33;font-size:13px;font-family:Manrope,sans-serif;color:{pill_fg}'>&#8593; {pill_txt}</div>"
            return f"<div style='font-size:14px;color:#a0a0a0;font-family:Manrope,sans-serif;margin-bottom:4px'>{label}</div><div style='font-size:36px;font-weight:700;color:var(--text-color,#1a1a1a);font-family:Manrope,sans-serif;line-height:1.1'>{value}</div>{pill}"

        m1,m2,m3,m4,m5 = st.columns(5)
        with m1: st.markdown(metric_card("Projects This Period",   f"{df[df['credit_tag'] != 'SKIPPED']['project'].nunique():,}"), unsafe_allow_html=True)
        with m2: st.markdown(metric_card("Hours This Period",      f"{hours_this_period:,.1f}"), unsafe_allow_html=True)
        with m3: st.markdown(metric_card("Utilization Credits",    f"{total_credit:,.1f}",    f"{credit_pct:.1%} of total hrs · {credit_label}", credit_color), unsafe_allow_html=True)
        with m4: st.markdown(metric_card("FF Project Overrun Hrs", f"{total_proj_overrun:,.1f}", f"{overrun_pct:.1%} of total hrs", "#ff4b4b"), unsafe_allow_html=True)
        with m5: st.markdown(metric_card("Admin Hrs",              f"{total_admin:,.1f}",     f"{admin_pct:.1%} of total hrs",    "#808495"), unsafe_allow_html=True)

        st.markdown("---")
        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            ["By Employee", "By Project", "ZCO Non-Billable", "Task Analysis", "Detail"]
        )

        with tab1:
            _ep = df[df["credit_tag"] != "SKIPPED"]
            emp_sum_ui = _ep.groupby(["employee","period"], as_index=False).agg(
                hours_this_period=("hours","sum"),
                credit_hrs=("credit_hrs","sum"),
                ff_overrun_hrs=("variance_hrs","sum"),
                admin_hrs=("hours", lambda x: df.loc[
                    (df["employee"].isin(_ep["employee"])) &
                    (df["billing_type"].str.lower()=="internal"), "hours"
                ].sum() if "billing_type" in df.columns else 0),
            ).sort_values(["employee","period"])
            # Build region lookup directly from df for UI context
            _emp_region_ui = df.dropna(subset=["region"]).groupby("employee")["region"].first().to_dict() if "region" in df.columns else {}
            emp_sum_ui["location"]   = emp_sum_ui["employee"].map(_emp_region_ui)
            emp_sum_ui["avail_hrs"]  = emp_sum_ui.apply(
                lambda r: get_avail_hours(r["location"], r["period"]) if r["location"] else None, axis=1)
            emp_sum_ui["util_pct"]   = emp_sum_ui.apply(
                lambda r: f"{r['credit_hrs']/r['avail_hrs']*100:.1f}%" if r["avail_hrs"] else "—", axis=1)
            display_cols = ["employee","location","period","avail_hrs",
                            "hours_this_period","credit_hrs","ff_overrun_hrs","util_pct"]
            st.dataframe(emp_sum_ui[[c for c in display_cols if c in emp_sum_ui.columns]],
                         use_container_width=True, hide_index=True)

        with tab2:
            proj_sum_ui = df[df["credit_tag"] != "SKIPPED"].groupby(
                ["project","project_type"], as_index=False
            ).agg(hours_this_period=("hours","sum"), credit_hrs=("credit_hrs","sum"),
                  ff_overrun_hrs=("variance_hrs","sum")).sort_values("project")
            st.dataframe(proj_sum_ui[["project","project_type","hours_this_period",
                         "credit_hrs","ff_overrun_hrs"]],
                         use_container_width=True, hide_index=True)

        with tab3:
            zco_df = df[df["credit_tag"] == "NON-BILLABLE"]
            if "task" in zco_df.columns and len(zco_df) > 0:
                zco_sum = zco_df.groupby(["task","employee","period"], as_index=False
                ).agg(hours=("hours","sum")).sort_values(["task","employee","period"])
                st.dataframe(zco_sum, use_container_width=True, hide_index=True)
            else:
                st.info("No ZCO Non-Billable entries in this dataset.")

        with tab4:
            ff_df = df[df["ff_task"].notna()] if "ff_task" in df.columns else pd.DataFrame()
            if len(ff_df) > 0:
                task_sum = ff_df.groupby(["ff_task","project_type"], as_index=False
                ).agg(hours=("hours","sum")).sort_values(["ff_task","project_type"])
                type_totals = ff_df.groupby("project_type")["hours"].sum()
                task_sum["pct_of_type"] = task_sum.apply(
                    lambda r: f"{r['hours']/type_totals.get(r['project_type'],1)*100:.1f}%", axis=1)
                st.dataframe(task_sum, use_container_width=True, hide_index=True)
            else:
                st.info("No Fixed Fee task data found. Check Billing Type and Task/Case columns.")

        with tab5:
            display_cols = ["employee","region","project","project_type","billing_type",
                            "hours_to_date","date","hours","credit_hrs","variance_hrs",
                            "previous_htd","credit_tag","notes"]
            existing = [c for c in display_cols if c in df.columns]
            st.dataframe(df[existing].head(100), use_container_width=True, hide_index=True)
            if len(df) > 100:
                st.caption(f"Showing first 100 of {len(df):,} rows. Full data in Excel download.")

        st.divider()

        # Download
        st.subheader("Download Report")
        with st.spinner("Building Excel file..."):
            excel_buf = build_excel(df, DEFAULT_SCOPE, consumed)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename  = f"utilization_report_{timestamp}.xlsx"

        st.download_button(
            label="⬇️ Download Excel Report",
            data=excel_buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
        st.caption(f"`{filename}` — 6 tabs: Processed Data · By Employee · By Project · "
                   f"ZCO Non-Billable · Task Analysis · Skipped Rows")


if __name__ == "__main__":
    main()
